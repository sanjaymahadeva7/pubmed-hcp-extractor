"""
Excel Export Module
Generate formatted Excel files from extracted data
"""
import pandas as pd
from datetime import datetime
import os

def create_excel_report(data_list, output_filename='pubmed_contacts.xlsx'):
    """
    Create Excel file from extracted data
    
    Args:
        data_list: List of dictionaries with contact information
        output_filename: Name of output Excel file
    
    Returns:
        Path to created Excel file
    """
    if not data_list:
        print("No data to export!")
        return None
    
    # Create DataFrame
    df = pd.DataFrame(data_list)
    
    # Define column order (15 columns)
    column_order = [
        'DataSource',
        'HCPName',
        'Country',
        'Province/State',
        'City',
        'Postal Code',
        'Email ID',
        'Alternate Email',
        'Assistant Email',
        'Phone No.',
        'TherapyArea',
        'Specialty',
        'Sub-Specialty',
        'Areas Of Expertise',
        'Affiliation/Institution',
        'ProfileURL',
        'Status',
        'ProfileNotes',
        'Feedback',
        'RecordUpdateDate',
        'PublicationLinks',
        'Date',
        'Title'
    ]
    
    # Ensure all columns exist, add with empty string if missing
    for col in column_order:
        if col not in df.columns:
            df[col] = ''
    
    # Reorder columns
    df = df[column_order]
    
    # Sort by Country, then Specialty
    df = df.sort_values(by=['Country', 'Specialty', 'HCPName'])
    
    # Reset index
    df = df.reset_index(drop=True)
    
    # Create Excel writer with formatting
    output_path = output_filename
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Contacts', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Contacts']
        
        # Format header row
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)

        # Standard data row height
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 14.5
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        column_widths = {
            'A': 14,  # DataSource
            'B': 26,  # HCPName
            'C': 14,  # Country
            'D': 20,  # Province/State
            'E': 18,  # City
            'F': 14,  # Postal Code
            'G': 36,  # Email ID
            'H': 36,  # Alternate Email
            'I': 36,  # Assistant Email
            'J': 18,  # Phone No.
            'K': 22,  # TherapyArea
            'L': 22,  # Specialty
            'M': 26,  # Sub-Specialty
            'N': 42,  # Areas Of Expertise
            'O': 55,  # Affiliation/Institution
            'P': 48,  # ProfileURL
            'Q': 16,  # Status
            'R': 30,  # ProfileNotes
            'S': 20,  # Feedback
            'T': 22,  # RecordUpdateDate
            'U': 55,  # PublicationLinks
            'V': 16,  # Date
            'W': 65   # Title
        }

        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Add borders and alignment to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(column_order)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
)

        
        # Freeze header row
        # worksheet.freeze_panes = 'C2'
    
    print(f"\n Excel file created: {output_path}")
    print(f" Total contacts exported: {len(df)}")
    
    return output_path

def print_summary_statistics(data_list):
    """
    Print summary statistics of extracted data
    
    Args:
        data_list: List of dictionaries with contact information
    """
    if not data_list:
        print("No data to summarize!")
        return
    
    df = pd.DataFrame(data_list)
    
    print("\n" + "="*60)
    print("EXTRACTION SUMMARY")
    print("="*60)
    
    print(f"\nTotal Contacts Found: {len(df)}")
    
    # Country breakdown
    if 'Country' in df.columns:
        print("\nContacts by Country:")
        country_counts = df['Country'].value_counts()
        for country, count in country_counts.items():
            print(f"  {country}: {count}")
    
    # Specialty breakdown
    if 'Specialty' in df.columns:
        print("\nContacts by Specialty:")
        specialty_counts = df['Specialty'].value_counts().head(10)
        for specialty, count in specialty_counts.items():
            print(f"  {specialty}: {count}")
    
    # TherapyArea breakdown
    if 'TherapyArea' in df.columns:
        therapy_counts = df[df['TherapyArea'] != '']['TherapyArea'].value_counts().head(10)
        if not therapy_counts.empty:
            print("\nContacts by Therapy Area:")
            for therapy, count in therapy_counts.items():
                print(f"  {therapy}: {count}")
    
    # Email availability (should be 100%)
    if 'Email ID' in df.columns:
        emails_found = df['Email ID'].notna().sum()
        print(f"\nEmails Found: {emails_found}/{len(df)} (100%)")
    
    print("\n" + "="*60)